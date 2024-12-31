import re
from copy import copy

from openpyxl import Workbook, load_workbook
import calendar
from datetime import datetime
print('test')
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from telebot.types import BotCommand, InlineKeyboardMarkup,  InlineKeyboardButton
import telebot
from config.auto_search_dir import data_config
import urllib3

# Отключаем предупреждения
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Создаем экземпляр бота
bot = telebot.TeleBot(data_config['my_telegram_bot']['bot_token'],
                      parse_mode='HTML')

class Main:
    def __init__(self):
        self.user_id = None
        self.register_commands()
        self.start_main()
        self.markup = []

    def register_commands(self):
        commands = [
            BotCommand("start", "В начало"),
        ]
        bot.set_my_commands(commands)

    def start_main(self):
        @bot.message_handler(commands=['start'])
        def handle_start_main(message):
            self.user_id = message.chat.id

            # Создаем клавиатуру с кнопками
            self.markup = InlineKeyboardMarkup()

            item2 = InlineKeyboardButton("Смены / подработки",
                                         callback_data='shifts')
            item3 = InlineKeyboardButton("Сотрудники",
                                         callback_data='employees')
            self.markup.add(item2, item3)

            bot.send_message(self.user_id,
                             "Используй кнопки для навигации. Чтобы вернуться на шаг назад, используй команду /back. "
                             "В начало /start",
                             reply_markup=self.markup)

        @bot.callback_query_handler(func=lambda call: True)
        def handle_query(call):
            if call.data == 'new_schedule':
                bot.send_message(call.message.chat.id,
                                 "Вы выбрали 'Новый график'.")
            elif call.data == 'shifts':
                bot.send_message(call.message.chat.id,
                                 "Вы выбрали 'Смены / подработки'.")
            elif call.data == 'employees':
                bot.send_message(call.message.chat.id,
                                 "Вы выбрали 'Сотрудники'.")


class Edit_chart(Main):
    def __init__(self):
        super().__init__()
        # список всех месяцев
        self.list_months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май',
                            'Июнь', 'Июль',
                            'Август', 'Сентябрь', 'Октябрь', 'Ноябрь',
                            'Декабрь']
        self.file = load_workbook('График работы.xlsx')  # загружаем файл
        self.weekdays = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']
        self.list_days_2 = {28: "AE", 29: "AF", 30: "AG", 31: "AH",
                            32: "AI", 33: "AJ", 34: "AK"}
        self.list_days = {"AE": 32, "AF": 33, "AG": 34, "AH": 35,
                          "AI": 36}  # 28,29, 30, 31 числа
        self.new_chart()  # запускаем основную функцию по созданию нового месяца
        self.first_weekday_next_month = []
        self.find_end_B_result = []
        self.source = []
        self.new_list = []
        self.result = []
        self.start_row = []
        self.start_col = []
        self.new_index = []
    def add_colls(self, colls, cell, i, new_value):
        for coll in colls:
            if coll in str(
                    cell):
                # присваиваем скопированному последнему столбцу переменную, где будет все храниться
                new_cell = self.new_list.cell(row=cell.row,
                                              column=cell.column + i,
                                              value=new_value)
                # если у этой ячейки есть стили , то также копируем их
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(
                        cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                    original_width = \
                        self.new_list.column_dimensions[
                            cell.column_letter].width
                    new_column_letter = self.new_list.cell(
                        row=cell.row,
                        column=cell.column + i).column_letter
                    self.new_list.column_dimensions[
                        new_column_letter].width = original_width

        # удаляем лишние звездочки до крайнего последнего столбца, лишние

    def remove(self, count_remove):
        self.new_list.delete_cols(
            self.list_days[self.list_days_2[self.result[0] - count_remove]],
            count_remove)
        self.new_list[
            f'{self.list_days_2[self.result[0] - count_remove]}1'] = '*'

    def copyRange(self, colls, count_add):
        # Определяем стиль шрифта
        font_style = Font(
            name='Calibri',
            charset=204,
            family=2,
            bold=False,
            italic=False,
            strike=False,
            outline=False,
            shadow=False,
            color='FF92D050',  # Цвет шрифта в формате RGB
            size=14.0,
        )

        # Определяем стиль заливки (если нужно)
        fill_style = PatternFill(
            start_color='FF92D050',  # Цвет фона в формате RGB
            end_color='FF92D050',
            fill_type='solid'
        )

        count_add_remove = abs(count_add)

        if count_add < 0:
            self.remove(count_add_remove)
        if count_add > 0:
            # получаем старый лист
            for i in range(1, count_add_remove + 1):
                for row in self.new_list.iter_rows():
                    for cell in row:  # получаем каждую ячейку
                        # добавляем дни для каждой новой ячейки
                        if cell.row == 4:
                            new_value = cell.value + i if isinstance(
                                cell.value, (int, float)) else cell.value
                        else:
                            new_value = cell.value
                        # если в новом месяце нужно добавить строки то выполняем , если нет, то исключение на удаление

                        self.add_colls(colls, cell, i, new_value)

            for col in range(25,
                             self.new_list.max_column):
                self.new_list.cell(row=1,
                                   column=col).value = None;

        # простовляем актуальные дни недели для нового месяца
        for col in range(4, self.new_list.max_column):
            start_index = self.first_weekday_next_month  # индекс первого для неделя для нового месяца
            for i in range(self.result[0] + self.result[
                1]):  # проходил полные дни для нового месяца
                # Вычисляем текущий индекс с помощью остатка от деления
                current_index = (start_index + i) % len(
                    self.weekdays)  # определяем новый день день недели для каждой итерации
                # изменяем значение дня недели на актуальный день
                self.new_list.cell(row=3, column=i + 4).value = self.weekdays[
                    current_index]

        # очищаем таблицу
        for row in range(5, self.find_end_B_result + 1):
            for col in range(4, self.new_list.max_column + 1):
                cell = self.new_list.cell(row=row, column=col)
                cell.value = None  # Устанавливаем значение на None
                cell.font = font_style  # Применяем стиль шрифта
                cell.fill = fill_style  # Применяем стиль заливки # Устанавливаем значение на None
                # if cell.has_style:
        if count_add > 0 or count_add < 0:
            # также изменяем ячейки для суммирования резульата
            for row in self.new_list.iter_rows(min_row=self.start_row,
                                               min_col=self.start_col):
                for cell in row:
                    if self.list_days_2[self.result[0]] in str(
                            cell.value):  # Получаем значения в текущей строке
                        text = str(cell.value).replace(
                            self.list_days_2[self.result[0]],
                            self.list_days_2[self.result[0] + self.result[1]])
                        cell.value = text

    def days_difference_current_next_month(self):
        # Получаем текущую дату
        current_date = datetime.now()

        # Определяем текущий год и месяц
        year = current_date.year
        month = current_date.month

        # month = 1

        # Функция для получения количества дней в месяце
        def days_in_month(year, month):
            return calendar.monthrange(year, month)[1]

        # Получаем количество дней в текущем месяце
        current_month_days = days_in_month(year, month)

        # Функция для получения следующего месяца
        def next_month(year, month):
            if month == 12:  # Декабрь
                return year + 1, 1  # Январь следующего года
            else:
                return year, month + 1  # Следующий месяц

        # Получаем следующий месяц
        next_year, next_month_value = next_month(year, month)

        # Получаем количество дней в следующем месяце
        next_month_days = days_in_month(next_year, next_month_value)
        # Получаем первый день недели следующего месяца
        # Получаем первый день недели следующего месяца
        self.first_weekday_next_month = \
            calendar.monthrange(next_year, next_month_value)[0]

        # Получаем название первого дня недели следующего месяца
        # Вычисляем разницу
        difference = [current_month_days,
                      (next_month_days - current_month_days)]

        return difference

    # новый график
    def new_chart(self):
        # Получаем лист самого крайнего месяца
        self.source = self.file.worksheets[-1]
        source_ = self.file.worksheets[0]

        self.file.remove(source_)
        # Создаем копию листа самого крайнего месяца
        self.new_list = self.file.copy_worksheet(self.source)

        # изменяем название страницы и заголовком на новый месяц
        for row in self.new_list.iter_rows():
            for cell in row:
                # получаем строку где расположен последний сотрудник
                search_end_B = str(cell)
                match = re.search(r"\.B(\d+)",
                                  search_end_B)  # ищем все строки , где есть B с цифрами и извлекаем их

                # если в найденной строке есть следующие строки
                if 'Сотрудник' == str(cell.value) and match:
                    # если есть то получаем только цифры, так как столбец будет незименяеммым
                    find_start_B_result = int(match.group(
                        1)) + 2  # отнимаем 2 так как через строчку будет расположен последний сотрудник

                # если в найденной строке есть следующие строки
                if 'работа//смена' == str(cell.value) and match:
                    # если есть то получаем только цифры, так как столбец будет незименяеммым
                    self.find_end_B_result = int(match.group(
                        1)) - 2  # отнимаем 2 так как через строчку будет расположен последний сотрудник
                    # если в найденной строке есть следующие строки
                if 'Итоги' in str(
                        cell.value):  # находим координаты для изменения ячеек для сумерования
                    match_2 = re.search(r"\.M(\d+)",
                                        str(cell))
                    # Начальные координаты
                    self.start_row = int(match_2.group(1)) + 4
                    self.start_col = 15

                    # если есть то получаем только цифры, так как столбец будет незименяеммым

                if cell.value in self.list_months:
                    find_index = self.list_months.index(cell.value)
                    self.new_index = (find_index + 1) % len(self.list_months)
                    self.new_list.title = f'{self.list_months[self.new_index]}'
                    # перед добавление нового месяца, удаляем месяц с таким же названием

                    # Изменяем названия прошлого месяца на новый
                    cell.value = f'{self.list_months[self.new_index]}'
                    # Проверяем, содержит ли ячейка "Итоги" и заменяем
                elif isinstance(cell.value, str) and "Итоги" in cell.value:
                    # Заменяем "Итоги (Декабрь)" на новый месяц
                    cell.value = f'Итоги ({self.list_months[self.new_index]})'

        self.result = self.days_difference_current_next_month()

        # разъединяЕм сначала ячейки чтобы добавить/удалить  новые столбцы
        self.new_list.unmerge_cells(
            f'B2:{self.list_days_2[self.result[0]]}2')

        # количество дней, которых не хватает в новом месяце
        count_add = self.result[1]
        result_1 = []
        if self.result[1] > 0:
            # находим крайнии строчки последнего дня , для последующего копирования и создания новых
            for i in range(1, self.find_end_B_result + 1):
                result_1.append(f'{(self.list_days_2[self.result[0]])}{i}')
        # функция для создания новой страницы и обработки дней, дней недель  и прочеее для нового месяца
        self.copyRange(result_1, count_add)
        # объединяем ячейки обратно где название месяца в основной таблице
        self.new_list.merge_cells(
            f'B2:{self.list_days_2[self.result[0] + self.result[1]]}2')
        # сохраняем итоговый вариант
        self.file.save('test1.xlsx')


# Запускаем бота
# if __name__ == "__main__":
Edit_chart()
# bot.polling(none_stop=True)
