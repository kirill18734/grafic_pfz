from openpyxl import load_workbook
from openpyxl.styles import *
from datetime import datetime
import calendar
from config.auto_search_dir import path_to_test1_json



# функция для получeния стилизации нужных ячеек
def get_font_style(color):
    value = None
    # Определяем стиль шрифта
    font_style = Font(
        name='Calibri',
        charset=204,
        family=2,
        bold=False,
        italic=False,
        strike=False,
        outline=False,
        shadow=False,
        color='FF92D050',  # Цвет шрифта в формате RGB
        size=14.0,
    )

    # Определяем стиль заливки
    fill_style = PatternFill(
        start_color='FF92D050',  # Цвет фона в формате RGB
        end_color='FF92D050',
        fill_type='solid'
    )
    number_format = 'General'
    # Определяем стиль границ
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Определяем стиль выравнивания
    alignment_style = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True
    )

    # Определяем защиту ячейки
    protection_style = Protection(
        locked=True,
        hidden=False
    )
    if color == 'green':
        return [thin_border, font_style, fill_style, number_format, protection_style, alignment_style, value]
    elif color == 'red':
        value = 1
        font_style.color = 'FF0000'
        fill_style.start_color = 'FF0000'
        fill_style.end_color = 'FF0000'
    elif color == 'blue':
        value = 1
        font_style.color = 'FF5B9BD5'
        fill_style.start_color = 'FF5B9BD5'
        fill_style.end_color = 'FF5B9BD5'
    elif color == 'orange':
        font_style.color = 'FF000000'
        fill_style.start_color = 'FFC000'
        fill_style.end_color = 'FFC000'
        alignment_style.horizontal = 'right'  # Устанавливаем выравнивание по правому краю
        alignment_style.vertical = 'center'  # Устанавливаем вертикальное выравнивание по центру (или 'top', 'bottom' по вашему выбору)
    return [thin_border, font_style, fill_style, number_format, protection_style, alignment_style, value]


class DataCharts:
    def __init__(self):
        self.list_months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май',
                            'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь',
                            'Декабрь']
        self.file = load_workbook(path_to_test1_json)  # загружаем файл
        # крайний месяц
        self.last_list = self.file.worksheets[-1]

    # получение активных сотрудников , за последний месяц
    def get_users(self):
        get_users = self.ineration_all_last_table(max_col=2, min_col=2, min_row=5)
        users = []
        for user in get_users:
            if user is not None:
                users.append(user)
            else:
                break
        return users

    def data_months(self):
        # Получаем текущую дату
        current_date = datetime.now()

        # Определяем текущий год и месяц
        year = current_date.year
        month = current_date.month

        # Функция для получения количества дней в месяце
        def days_in_month(nex_year, nex_month):
            return calendar.monthrange(nex_year, nex_month)[1]

        # Получаем количество дней в текущем месяце
        current_month_days = days_in_month(year, month)

        # Функция для получения следующего месяца
        def next_month(nex_year, nex_month):
            if nex_month == 12:  # Декабрь
                return nex_year + 1, 1  # Январь следующего года
            else:
                return nex_year, nex_month + 1  # Следующий месяц

        # Получаем следующий месяц
        next_year, next_month_value = next_month(year, month)

        # Получаем количество дней в следующем месяце
        next_month_days = days_in_month(next_year, next_month_value)

        # Получаем первый день недели следующего месяца
        first_weekday_next_month = calendar.monthrange(next_year, next_month_value)[0]
        new_index = month % len(self.list_months)
        # количество дней в следующем месяце, разницу в днях,  индекс текущего месяца, индекс следующего месяца
        difference = [current_month_days,
                      (next_month_days - current_month_days) + 3, month -1 , new_index, first_weekday_next_month]
        return difference

    # функция для получения нужные данных
    def ineration_all_last_table(self, min_row=None, min_col=None, max_col=None, max_row=None):
        result = []
        for row in self.last_list.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            for cell in row:
                result.append(cell.value)
        return result