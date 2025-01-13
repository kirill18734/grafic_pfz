from openpyxl.drawing import fill

from config.auto_search_dir import path_to_test1_json
from edit_charts.data_file import DataCharts, get_font_style
from send_to_telegram_email.send_to_TG_email import test_mode


class Editsmens:
    def __init__(self):
        self.file = None
        self.table = DataCharts()

    def smens(self, month, user):
        self.file = self.table.file[month]
        # получаем номер строки где наш пользователь
        find_row = [cell.row for row in self.file.iter_rows(
            max_row=len(self.table.get_users()) + 4) for cell in row if
                    user in str(cell.value)]
        result = {}
        count = 1
        for row in self.file.iter_rows(min_col=4, max_row=find_row[0],
                                       min_row=find_row[0]):
            for cell in row:
                if cell.fill.start_color.rgb == 'FF00B0F0':
                    result[f'{count}i'] = cell.value
                else:
                    result[count] = cell.value
                count += 1

        return result

    def get_days(self, month):
        self.file = self.table.file[month]
        result = {}
        count = 1
        for row in self.file.iter_rows(min_col=4, max_row=3, min_row=3):
            for cell in row:
                result[count] = cell.value
                count += 1
        return result

    def edit_smens(self, month, user, new_smens):
        self.file = self.table.file[month]

        # Получаем номер строки, где наш пользователь
        find_row = [cell.row for row in self.file.iter_rows(
            max_row=len(self.table.get_users()) + 4) for cell in row if
                    user in str(cell.value)]

        if not find_row:
            print("Пользователь не найден.")
            return

        # Предполагаем, что мы работаем только с первой найденной строкой
        target_row = find_row[0]

        # Преобразуем генератор в список, чтобы получить доступ к ячейкам
        row_cells = list(self.file.iter_rows(min_col=4, max_row=target_row,
                                             min_row=target_row))[0]

        # Перебираем ячейки в целевой строке
        for count, cell in enumerate(row_cells, start=1):
            key = count if count in new_smens else None

            if key is not None:
                value = new_smens[key]
                color = None

                if value is None:
                    color = 'green'
                elif value == 1:
                    color = 'red'
                elif value > 1:
                    color = 'orange'
            else:
                # Проверяем, если ключ строка с 'i'
                str_key = str(count) + 'i'
                if str_key in new_smens:
                    value = new_smens[str_key]
                    if value == 1:
                        color = 'blue'
                    else:
                        continue  # Если значение не 1, пропускаем

            # Обновляем ячейку
            if color:
                cell.value = value
                cell.border = get_font_style(color)[0]
                cell.font = get_font_style(color)[1]
                cell.fill = get_font_style(color)[2]
                cell.number_format = get_font_style(color)[3]
                cell.protection = get_font_style(color)[4]
                cell.alignment = get_font_style(color)[5]
                if color != 'orange':
                    cell.value = get_font_style(color)[6]

        self.table.file.save(path_to_test1_json)
        self.table.file.close()

#
# test = Editsmens()
# print(test.edit_smens('Январь', 'Кирилл', {1: None, 2: None, 3: 3, 4: None, 5: None, 6: None, 7: 1, 8: None, 9: None, 10: None, 11: None, 12: None, 13: None, 14: None, 15: None, 16: None, 17: None, 18: None, 19: None, 20: None, 21: None, '22i': 1, 23: 7, 24: None, 25: None, 26: None, 27: None, 28: None, 29: None, 30: None, 31: 1}))