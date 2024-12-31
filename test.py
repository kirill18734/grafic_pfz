import openpyxl
from copy import copy
from openpyxl.reader.excel import load_workbook

# Создаем новую рабочую книгу и добавляем лист по умолчанию
workbook = load_workbook('test1.xlsx')
default_sheet = workbook.active
default_sheet.title = 'Январь1'

# Создаем новый лист
sheetName = 'NewSheet'
new_sheet = workbook.create_sheet(sheetName)

# Копируем данные и стили из листа по умолчанию в новый лист
for row in default_sheet.iter_rows():
    for cell in row:
        print(cell)
        # Используем cell.row и cell.column вместо cell.row и cell.col_idx
        new_cell = new_sheet.cell(row=cell.row, column=cell.column +1, value=cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
        print(default_sheet.iter_rows())

# Сохраняем рабочую книгу
# workbook.save('test_workbook.xlsx')

print("Рабочая книга успешно создана и сохранена.")
